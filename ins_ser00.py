import time
import pandas as pd
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
# Configurar caminho do ChromeDriver e opções do navegador
chrome_options = Options()
chrome_options.add_argument("--start-maximized")

service = Service(r'C:\Users\matheus.paixao\Downloads\Python test\chromedriver-win64\chromedriver.exe')
driver = webdriver.Chrome(service=service, options=chrome_options)

# Abrir a URL inicial
driver.get('https://appweb1.antt.gov.br/sar/Site/Inscricao/Serasa/SelecionarPreLoteInclusaoSerasa.aspx')

def login_sifama():
    try:
        usuario = WebDriverWait(driver, 0).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_TextBoxUsuario"]'))
        )
        senha = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_TextBoxSenha"]')

        usuario.send_keys("matheus.paixao")
        senha.send_keys("#Omnes22")
        senha.send_keys(Keys.RETURN)

        # Aguardar e atualizar
        time.sleep(5)
        driver.get('https://appweb1.antt.gov.br/sar/Site/Inscricao/Serasa/SelecionarPreLoteInclusaoSerasa.aspx')
    except Exception as e:
        print(f"Erro ao tentar fazer login: {e}")
        driver.quit()

# Fazer login no sistema
login_sifama()

# Carregar planilha existente com os autos de infração da aba "PRODUÇÃO"
excel_path = r'C:\Users\matheus.paixao\Downloads\Python test\processos aptos.xlsx'
df = pd.read_excel(excel_path, sheet_name='PRODUÇÃO')
#Lista para armazenar autos não encontrados
nao_encontrados = []
# Função para verificar e incluir no Serasa
def verificar_incluir_auto(auto_numero):
    try:
        # Localizar e preencher o campo do auto de infração
        campo_auto = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="Corpo_txbAutoInfracao"]'))
        )
        campo_auto.clear()
        campo_auto.send_keys(auto_numero)

        # Clicar no botão pesquisar
        botao_pesquisar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Corpo_btnPesquisar"]'))
        )
        botao_pesquisar.click()
        #Clicar na caixa de seleção
        
        Mark_box = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Corpo_gdvAutoInfracao_ckSelecionar_0"]'))
        )
        if not Mark_box.is_selected():
            driver.execute_script("arguments[0].click();", Mark_box)

        print(f"Auto {auto_numero} encontrado")
        return True
    except Exception as e:
        print(f"Nenhum registro encontrado para {auto_numero}: {e}")
        return False

# Iterar sobre os números de auto na coluna 'IDENTIFICADOR DE DÉBITO' da planilha
for auto_numero in df['IDENTIFICADOR DE DÉBITO']:
    verificar_incluir_auto(auto_numero)
    
# Iterar sobre os números de auto na coluna 'IDENTIFICADOR DE DÉBITO' da planilha, com limite de 500 autos
contador = 0
for auto_numero in df['IDENTIFICADOR DE DÉBITO']:
    if contador >= 9:
        break
    if not verificar_incluir_auto(auto_numero):
        nao_encontrados.append(auto_numero)
    contador += 1

# Criar novo Excel com os autos não encontrados, se necessário
if nao_encontrados:
    wb = Workbook()
    ws = wb.active
    ws.append(['Auto de Infração Não Encontrado'])

    for auto in nao_encontrados:
        ws.append([auto])

    wb.save(r'C:\Users\matheus.paixao\Downloads\Python test\autos_nao_encontrados.xlsx')
# n Encerrar o navegador
# driver.quit()

--------------------------------------------------------------------//////-------------------------------------------------------------------------------------------------------------

import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException

# Configurar caminho do ChromeDriver e opções do navegador
chrome_options = Options()
chrome_options.add_argument("--start-maximized")

service = Service(r'C:\Users\matheus.paixao\Downloads\Python test\chromedriver-win64\chromedriver.exe')
driver = webdriver.Chrome(service=service, options=chrome_options)

# Abrir a URL inicial
driver.get('https://appweb1.antt.gov.br/sar/Site/Inscricao/Serasa/SelecionarPreLoteInclusaoSerasa.aspx')

def login_sifama():
    try:
        usuario = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_TextBoxUsuario"]'))
        )
        senha = driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_ContentPlaceHolderCorpo_TextBoxSenha"]')

        usuario.send_keys("matheus.paixao")
        senha.send_keys("#Omnes22")
        senha.send_keys(Keys.RETURN)

        # Aguardar e atualizar
        time.sleep(5)
        driver.get('https://appweb1.antt.gov.br/sar/Site/Inscricao/Serasa/SelecionarPreLoteInclusaoSerasa.aspx')
    except Exception as e:
        print(f"Erro ao tentar fazer login: {e}")
        driver.quit()

# Fazer login no sistema
login_sifama()

# Carregar planilha existente com os autos de infração da aba "PRODUÇÃO"
excel_path = r'C:\Users\matheus.paixao\Downloads\Python test\processos aptos.xlsx'
wb = load_workbook(excel_path)
ws = wb['PRODUÇÃO']
df = pd.read_excel(excel_path, sheet_name='PRODUÇÃO')

# Cores para preenchimento
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
# Função para tratar o erro de serviço usando ESC
def tratar_erro_servico():
    try:
        # Aguardar a mensagem de erro
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, 'divModalLabel'))
        )
        # Pressionar a tecla ESC para fechar o modal
        webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        print("Erro de serviço tratado com ESC. Tentando novamente...")
        return True
    except TimeoutException:
        return False
# Função para marcar a caixa de seleção com retry
def marcar_caixa_selecao(driver, xpath):
    try:
        Mark_box = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        time.sleep(1)
        if not Mark_box.is_selected():
            driver.execute_script("arguments[0].click();", Mark_box)
        return True
    except (StaleElementReferenceException, TimeoutException):
        return False

# Função para verificar e incluir no Serasa
def verificar_incluir_auto(auto_numero, row, tentativas=0):
    try:
        time.sleep(1)
        # Localizar e preencher o campo do auto de infração
        campo_auto = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="Corpo_txbAutoInfracao"]'))
        )
        campo_auto.clear()
        campo_auto.send_keys(auto_numero)

        # Clicar no botão pesquisar
        time.sleep(1)
        botao_pesquisar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Corpo_btnPesquisar"]'))
        )
        botao_pesquisar.click()

        # Verificar se a mensagem de "Nenhum registro encontrado" está presente
        try:
            time.sleep(2)
            mensagem_erro = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="Corpo_gdvAutoInfracao"]/tbody/tr/td'))
            )
            if mensagem_erro.text == "Nenhum registro encontrado.":
                print(f"Nenhum registro encontrado para {auto_numero}")
                ws.cell(row=row, column=1).fill = red_fill
                return "not_found"
        except NoSuchElementException:
            pass

        # Clicar na caixa de seleção com retry
        time.sleep(1)
        if marcar_caixa_selecao(driver, '//*[@id="Corpo_gdvAutoInfracao_ckSelecionar_0"]'):
            print(f"Auto {auto_numero} encontrado")
            ws.cell(row=row, column=1).fill = green_fill  # Corrigido
            return True
        else:
            time.sleep(1)
            print(f"Falha ao tentar marcar a caixa de seleção para {auto_numero}.")
            ws.cell(row=row, column=1).fill = red_fill  # Corrigido
            return False
        
    except TimeoutException:
        time.sleep(2)
        print(f"Timeout ao tentar verificar e incluir o auto {auto_numero}.")
        ws.cell(row=row, column=1).fill = red_fill  # Corrigido
        return False
    except StaleElementReferenceException:
        time.sleep(4)
        print(f"Stale Element Reference ao tentar verificar e incluir o auto {auto_numero}.")
        ws.cell(row=row, column=1).fill = red_fill  # Corrigido
        return False
    except Exception as e:
        if "Falha na execução do serviço" in str(e) and tentativas < 3:
            if tratar_erro_servico():
                return verificar_incluir_auto(auto_numero, row, tentativas + 1)
        time.sleep(4)
        print(f"Nenhum registro encontrado para {auto_numero}: {e}")
        ws.cell(row=row, column=1).fill = red_fill
    
        return "not_found"
    
# Registrar o tempo de início
inicio = time.time()
# Iterar sobre os números de auto na coluna 'IDENTIFICADOR DE DÉBITO' da planilha, com limite de 200 autos
contador = 0
for index, row_data in df.iterrows():
    if contador >= 500:
        break
    auto_numero = row_data['IDENTIFICADOR DE DÉBITO']
    # Verificar se o auto pode ser marcado
    verificar_incluir_auto(auto_numero, index + 2)  # +2 para ajustar o índice do DataFrame ao índice do Excel
    contador += 1

# Salvar as alterações na planilha
wb.save(excel_path)
# Registrar o tempo de fim
fim = time.time()
tempo_total = fim - inicio
# Exibir o resumo
print("Processo concluído.")
print(f"Tempo total: {tempo_total:.2f} segundos")

# Encerrar o navegador
#driver.quit()


